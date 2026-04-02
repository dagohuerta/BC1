
"""
retail_roi_model.py

Python implementation of the main calculation flow used by the Excel workbook:
USD_Retail Insight Unified Financial Model_FY26_CS_Coppel_V3.xlsm

Scope covered:
1) Financial input extraction
2) Forward financials by quarter (5 years / 20 quarters)
3) Module-level benefits and investment timing
4) Total quarterly and annual ROI output
5) Incremental P&L / cash flow
6) High-level pro forma outputs
7) NPV / IRR / payback

Important design choice:
- The code is workbook-driven. It reads the workbook structure and exact input cells,
  but the calculations are reproduced in Python rather than reading the Excel outputs.
- The formulas implemented here follow the main logic from:
  Financial Input
  Benefit Input
  Value Benefit
  Adoption Input
  Investment Input
  Timeline Input
  Forward Financials
  ROI Output - Modular
  ROI Output - Total
  P&L and Cash Flow
  Proforma Financials

Notes:
- This implementation focuses on the core finance engine used by the workbook.
- It does not recreate Excel formatting, chart sheets, or every diagnostic row.
- Category-level rollups are calculated directly from module results instead of
  reconstructing the intermediate "ROI Output - Category" sheet.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


QUARTERS_PER_YEAR = 4
TOTAL_YEARS = 5
TOTAL_QUARTERS = QUARTERS_PER_YEAR * TOTAL_YEARS
QUARTER_LABELS = ["Q1", "Q2", "Q3", "Q4"] * TOTAL_YEARS


def safe_num(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return float(value)
    try:
        return float(value)
    except Exception:
        return default


def chunk(values: List[float], size: int) -> List[List[float]]:
    return [values[i:i + size] for i in range(0, len(values), size)]


def annualize(quarterly: List[float]) -> List[float]:
    return [sum(x) for x in chunk(quarterly, QUARTERS_PER_YEAR)]


def cumulative(values: List[float]) -> List[float]:
    out: List[float] = []
    running = 0.0
    for v in values:
        running += v
        out.append(running)
    return out


def npv(rate: float, cashflows: List[float]) -> float:
    return sum(cf / ((1.0 + rate) ** i) for i, cf in enumerate(cashflows, start=1))


def irr(cashflows: List[float], low: float = -0.9999, high: float = 10.0, iterations: int = 200) -> Optional[float]:
    """
    Bisection IRR solver on annual cash flows.
    Returns None when sign change is not present.
    """
    def f(r: float) -> float:
        return npv(r, cashflows)

    f_low = f(low)
    f_high = f(high)

    if math.isnan(f_low) or math.isnan(f_high):
        return None
    if f_low == 0:
        return low
    if f_high == 0:
        return high
    if f_low * f_high > 0:
        return None

    a, b = low, high
    for _ in range(iterations):
        mid = (a + b) / 2.0
        f_mid = f(mid)
        if abs(f_mid) < 1e-9:
            return mid
        if f_low * f_mid < 0:
            b = mid
            f_high = f_mid
        else:
            a = mid
            f_low = f_mid
    return (a + b) / 2.0


def payback_period(cumulative_net_benefit_by_year: List[float]) -> Optional[str]:
    prev_value = 0.0
    for idx, value in enumerate(cumulative_net_benefit_by_year, start=1):
        if value > 0:
            if idx == 1:
                # Interpolación en el primer año (asumiendo inversión inicial en t=0)
                # Si el valor acumulado al final del año 1 es positivo, el payback ocurrió dentro del año 1.
                # Como no tenemos el valor en t=0 (inversión), asumimos que el beneficio es lineal.
                # Inversión = value - beneficio_año_1. Pero no tenemos beneficio_año_1 directamente aquí.
                # Supondremos que prev_value es el valor en t=0 (negativo).
                # Pero en la lista pasada, prev_value empieza en 0.
                return "~ 12 meses"
            
            # Interpolación: meses = (años_previos * 12) + (falta_para_cero / beneficio_este_año * 12)
            beneficio_año = value - prev_value
            if beneficio_año > 0:
                meses_adicionales = (abs(prev_value) / beneficio_año) * 12
                total_meses = int((idx - 1) * 12 + meses_adicionales)
                return f"{total_meses} meses"
            return f"{idx * 12} meses"
        prev_value = value
    return None


def quarter_factor(quarter_index: int, quarter_mix: List[float]) -> float:
    return quarter_mix[quarter_index % 4]


def lookup_adoption_value(seq: int, sequence_axis: List[int], row_values: List[float]) -> float:
    """
    Replicates the workbook HLOOKUP behavior on Adoption Input where:
    - row 8 is the sequence axis: 0,1,2,3...
    - subsequent rows contain values by live-quarter sequence
    """
    if not sequence_axis:
        return 0.0
    if seq <= sequence_axis[0]:
        return safe_num(row_values[0])
    for i, axis_value in enumerate(sequence_axis):
        if seq == axis_value:
            return safe_num(row_values[i])
        if seq < axis_value:
            return safe_num(row_values[i - 1])
    return safe_num(row_values[-1])


@dataclass
class FinancialInputs:
    client_name: str
    retailer_type: str

    # Income statement / balance sheet
    net_revenue: float
    cogs: float
    gross_profit: float
    sga: float
    dep_amort_etc: float
    operating_income: float
    interest: float
    taxes: float
    net_income: float

    cash: float
    accounts_receivable: float
    total_inventory: float
    other_current_assets: float
    gross_working_capital: float
    property_equipment: float
    goodwill: float
    other_assets: float
    total_assets: float
    accounts_payable: float
    other_current_liabilities: float
    current_liabilities: float
    long_term_debt: float
    other_long_term_liabilities: float
    total_liabilities: float
    total_equity: float

    # Drivers / assumptions
    annual_inventory_turns: float
    tax_rate: float
    inventory_carrying_cost_of_capital: float
    annual_revenue_growth_rate: float
    annual_cogs_growth_rate: float
    annual_sga_growth_rate: float
    npv_discount_rate: float
    interest_rate_on_cash_flow: float
    ecommerce_pct_of_business: float
    ecommerce_gross_margin: float

    # Quarterly sales pattern
    quarterly_revenue: List[float]
    quarterly_mix: List[float]

    # Business breakdown block (rows 85:94)
    business_breakdown_names: List[str] = field(default_factory=list)
    business_breakdown_revenue_pct: List[float] = field(default_factory=list)
    business_breakdown_revenue: List[float] = field(default_factory=list)
    business_breakdown_margin_pct: List[float] = field(default_factory=list)
    business_breakdown_inventory_pct: List[float] = field(default_factory=list)
    business_breakdown_inventory: List[float] = field(default_factory=list)
    business_breakdown_full_price_pct: List[float] = field(default_factory=list)
    business_breakdown_discounted_price_pct: List[float] = field(default_factory=list)


@dataclass
class ModuleAssumptions:
    module_name: str
    admin_row: int
    selected: bool

    benefit_row: Optional[int]
    value_row: Optional[int]
    adoption_base_row: Optional[int]
    investment_row: Optional[int]
    roi_start_row: Optional[int]
    timeline_status_row: Optional[int]
    timeline_sequence_row: Optional[int]

    # Workbook control cells from ROI Output - Modular
    sales_business_impacted_pct: float = 0.0     # F[row_start]
    inventory_business_impacted_pct: float = 0.0 # F[row_start+1]
    impacted_margin_pct: float = 0.0             # J[row_start]
    labor_pct_of_sga: float = 0.0                # J[row_start+1]

    # Benefit Input (conservative columns used by formulas)
    benefit_increased_sales_pct: float = 0.0     # Benefit Input C
    benefit_inventory_reduction_pct: float = 0.0 # Benefit Input F
    benefit_margin_improvement_pct: float = 0.0  # Benefit Input I
    benefit_labor_reduction_pct: float = 0.0     # Benefit Input L
    benefit_logistics_reduction_pct: float = 0.0 # Benefit Input O

    # Value Benefit inputs / subtotals used by formulas
    vb_profit_sales_subtotal: float = 0.0        # Value Benefit I
    vb_one_time_inventory: float = 0.0           # Value Benefit J
    vb_inventory_reduction: float = 0.0          # Value Benefit K
    vb_margin_improvement: float = 0.0           # Value Benefit L
    vb_labor_sga_subtotal: float = 0.0           # Value Benefit T
    vb_logistics: float = 0.0                    # Value Benefit U
    vb_carrying_cost_benchmark: float = 0.0      # Value Benefit J41 (global benchmark used in workbook)

    # Adoption rows from Adoption Input
    adoption_sequence_axis: List[int] = field(default_factory=list)  # row 8, C:S
    adoption_recurrent_curve: List[float] = field(default_factory=list)   # base row, C:S
    adoption_one_time_curve: List[float] = field(default_factory=list)    # base row + 1, C:S

    # Timeline input
    timeline_statuses: List[int] = field(default_factory=list)      # row status, D:W
    timeline_live_sequence: List[int] = field(default_factory=list) # row sequence, D:W

    # Investment Input
    software_fee: float = 0.0             # AG
    software_maintenance: float = 0.0     # AH
    hardware_capital: float = 0.0         # AI
    hardware_maintenance: float = 0.0     # AJ
    hosting_fees: float = 0.0             # AK
    oracle_services: float = 0.0          # J
    third_party_services: float = 0.0     # R
    client_services: float = 0.0          # Z
    financing_quarters: int = 0           # Timeline Input Z10
    implementation_quarters: int = 0      # COUNTIF(status row, 1)

    # Optional metadata
    assessment_type: Optional[str] = None


@dataclass
class ModuleQuarterlyResult:
    module_name: str
    selected: bool
    profit_from_increased_sales: List[float]
    inventory_reductions: List[float]
    one_time_inventory_reduction: List[float]
    carrying_cost_reduction: List[float]
    margin_improvement: List[float]
    labor_sga_reduction: List[float]
    logistics_reduction: List[float]
    total_estimated_benefit: List[float]

    software_fees: List[float]
    software_maintenance: List[float]
    oracle_services: List[float]
    client_services: List[float]
    third_party_integrator_services: List[float]
    hardware_capital_costs: List[float]
    hardware_maintenance: List[float]
    hosting_fees: List[float]
    total_investment: List[float]

    cumulative_estimated_benefit: List[float]
    cumulative_investment: List[float]
    cumulative_net_benefit: List[float]
    cumulative_roi: List[float]


@dataclass
class ModelOutputs:
    forward_financials: Dict[str, List[float]]
    module_results: List[ModuleQuarterlyResult]
    quarterly_total: Dict[str, List[float]]
    annual_total: Dict[str, List[float]]
    pnl_cashflow: Dict[str, List[float]]
    proforma: Dict[str, Dict[str, List[float] | float]]
    summary_metrics: Dict[str, Any]


class WorkbookDrivenRetailROIModel:
    def __init__(self, workbook_path: str | Path):
        self.workbook_path = str(workbook_path)
        self.wb_formula = openpyxl.load_workbook(self.workbook_path, data_only=False, keep_vba=True)
        self.wb_value = openpyxl.load_workbook(self.workbook_path, data_only=True, keep_vba=True)

    def _defined_name_value(self, name: str) -> float:
        dn = self.wb_formula.defined_names.get(name)
        if dn is None:
            return 0.0
        dest = dn.attr_text
        m = re.match(r"'?([^']+)'?!\$?([A-Z]+)\$?(\d+)", dest)
        if not m:
            return 0.0
        sheet_name, col, row = m.group(1), m.group(2), int(m.group(3))
        return safe_num(self.wb_value[sheet_name][f"{col}{row}"].value)

    def load_financial_inputs(self) -> FinancialInputs:
        ws = self.wb_value["Financial Input"]
        module_ws = self.wb_value["Module Selection"]

        client_name = module_ws["C8"].value or ""
        retailer_type = module_ws["E8"].value or ""

        bb_names = [ws[f"D86"].value, ws[f"E86"].value, ws[f"F86"].value]
        bb_rev_pct = [safe_num(ws[f"{col}87"].value) for col in ["D", "E", "F"]]
        bb_rev = [safe_num(ws[f"{col}88"].value) for col in ["D", "E", "F"]]
        bb_margin_pct = [safe_num(ws[f"{col}89"].value) for col in ["D", "E", "F"]]
        bb_inventory_pct = [safe_num(ws[f"{col}91"].value) for col in ["D", "E", "F"]]
        bb_inventory = [safe_num(ws[f"{col}92"].value) for col in ["D", "E", "F"]]
        bb_full_price_pct = [safe_num(ws[f"{col}93"].value) for col in ["D", "E", "F"]]
        bb_discount_pct = [safe_num(ws[f"{col}94"].value) for col in ["D", "E", "F"]]

        return FinancialInputs(
            client_name=str(client_name),
            retailer_type=str(retailer_type),

            net_revenue=safe_num(ws["D15"].value),
            cogs=safe_num(ws["D16"].value),
            gross_profit=safe_num(ws["D17"].value),
            sga=safe_num(ws["D18"].value),
            dep_amort_etc=safe_num(ws["D19"].value),
            operating_income=safe_num(ws["D20"].value),
            interest=safe_num(ws["D21"].value),
            taxes=safe_num(ws["D22"].value),
            net_income=safe_num(ws["D23"].value),

            cash=safe_num(ws["D27"].value),
            accounts_receivable=safe_num(ws["D28"].value),
            total_inventory=safe_num(ws["D29"].value),
            other_current_assets=safe_num(ws["D30"].value),
            gross_working_capital=safe_num(ws["D31"].value),
            property_equipment=safe_num(ws["D32"].value),
            goodwill=safe_num(ws["D33"].value),
            other_assets=safe_num(ws["D34"].value),
            total_assets=safe_num(ws["D35"].value),
            accounts_payable=safe_num(ws["D36"].value),
            other_current_liabilities=safe_num(ws["D37"].value),
            current_liabilities=safe_num(ws["D38"].value),
            long_term_debt=safe_num(ws["D39"].value),
            other_long_term_liabilities=safe_num(ws["D40"].value),
            total_liabilities=safe_num(ws["D41"].value),
            total_equity=safe_num(ws["D42"].value),

            annual_inventory_turns=safe_num(ws["D45"].value),
            tax_rate=safe_num(ws["D62"].value),
            inventory_carrying_cost_of_capital=safe_num(ws["D63"].value),
            annual_revenue_growth_rate=safe_num(ws["D64"].value),
            annual_cogs_growth_rate=safe_num(ws["D65"].value),
            annual_sga_growth_rate=safe_num(ws["D66"].value),
            npv_discount_rate=safe_num(ws["D67"].value),
            interest_rate_on_cash_flow=safe_num(ws["D68"].value),
            ecommerce_pct_of_business=safe_num(ws["D70"].value),
            ecommerce_gross_margin=safe_num(ws["D71"].value),

            quarterly_revenue=[safe_num(ws[f"D{r}"].value) for r in range(55, 59)],
            quarterly_mix=[safe_num(ws[f"E{r}"].value) for r in range(55, 59)],

            business_breakdown_names=[str(x) if x is not None else "" for x in bb_names],
            business_breakdown_revenue_pct=bb_rev_pct,
            business_breakdown_revenue=bb_rev,
            business_breakdown_margin_pct=bb_margin_pct,
            business_breakdown_inventory_pct=bb_inventory_pct,
            business_breakdown_inventory=bb_inventory,
            business_breakdown_full_price_pct=bb_full_price_pct,
            business_breakdown_discounted_price_pct=bb_discount_pct,
        )

    def _admin_row_from_formula(self, formula: str) -> Optional[int]:
        if not isinstance(formula, str):
            return None
        m = re.search(r"Admin - Sheet_Row'!B(\d+)", formula)
        return int(m.group(1)) if m else None

    def _roi_start_rows(self) -> Dict[int, int]:
        ws = self.wb_formula["ROI Output - Modular"]
        out: Dict[int, int] = {}
        for r in range(1, ws.max_row + 1):
            admin_row = self._admin_row_from_formula(ws[f"B{r}"].value)
            if admin_row is not None:
                out[admin_row] = r
        return out

    def load_modules(self) -> List[ModuleAssumptions]:
        admin = self.wb_value["Admin - Sheet_Row"]
        module_selection = self.wb_value["Module Selection"]
        benefit = self.wb_value["Benefit Input"]
        value_benefit = self.wb_value["Value Benefit"]
        adoption = self.wb_value["Adoption Input"]
        investment = self.wb_value["Investment Input"]
        timeline = self.wb_value["Timeline Input"]
        roi_modular = self.wb_value["ROI Output - Modular"]
        roi_start_map = self._roi_start_rows()

        modules: List[ModuleAssumptions] = []
        global_vb_carry = safe_num(value_benefit["J41"].value)

        for admin_row in range(12, 69):
            name = admin[f"B{admin_row}"].value
            if not name:
                continue

            selected = bool(module_selection[f"D{admin_row + 8}"].value) if admin_row + 8 <= module_selection.max_row else False
            assessment_type = module_selection[f"E{admin_row + 8}"].value if admin_row + 8 <= module_selection.max_row else None

            benefit_row = admin[f"C{admin_row}"].value
            value_row = admin[f"E{admin_row}"].value
            adoption_row = admin[f"G{admin_row}"].value
            investment_row = None

            # Investment sheet starts module list on row 10 aligned to admin row 12
            investment_row = admin_row - 2 if admin_row >= 12 else None

            roi_start = roi_start_map.get(admin_row)

            # Category headers / placeholders may not have row mappings
            if isinstance(benefit_row, str) and benefit_row.startswith("="):
                benefit_row = None
            if isinstance(value_row, str) and value_row.startswith("="):
                value_row = None
            if isinstance(adoption_row, str) and adoption_row.startswith("="):
                adoption_row = None

            timeline_status_row = None
            timeline_sequence_row = None
            if adoption_row is not None:
                # Workbook pattern:
                # adoption row 10 -> timeline status row 18 and sequence row 19
                # every additional module increments by 4 in adoption and by 3 in timeline
                offset = int(benefit_row) - 10 if benefit_row is not None else 0
                timeline_status_row = 18 + (offset * 3)
                timeline_sequence_row = timeline_status_row + 1

            m = ModuleAssumptions(
                module_name=str(name),
                admin_row=admin_row,
                selected=selected,
                benefit_row=int(benefit_row) if benefit_row is not None else None,
                value_row=int(value_row) if value_row is not None else None,
                adoption_base_row=int(adoption_row) if adoption_row is not None else None,
                investment_row=int(investment_row) if investment_row is not None else None,
                roi_start_row=roi_start,
                timeline_status_row=timeline_status_row,
                timeline_sequence_row=timeline_sequence_row,
                vb_carrying_cost_benchmark=global_vb_carry,
                assessment_type=str(assessment_type) if assessment_type is not None else None,
            )

            if m.benefit_row is not None:
                m.benefit_increased_sales_pct = safe_num(benefit[f"C{m.benefit_row}"].value)
                m.benefit_inventory_reduction_pct = safe_num(benefit[f"F{m.benefit_row}"].value)
                m.benefit_margin_improvement_pct = safe_num(benefit[f"I{m.benefit_row}"].value)
                m.benefit_labor_reduction_pct = safe_num(benefit[f"L{m.benefit_row}"].value)
                m.benefit_logistics_reduction_pct = safe_num(benefit[f"O{m.benefit_row}"].value)

            if m.value_row is not None:
                m.vb_profit_sales_subtotal = safe_num(value_benefit[f"I{m.value_row}"].value)
                m.vb_one_time_inventory = safe_num(value_benefit[f"J{m.value_row}"].value)
                m.vb_inventory_reduction = safe_num(value_benefit[f"K{m.value_row}"].value)
                m.vb_margin_improvement = safe_num(value_benefit[f"L{m.value_row}"].value)
                m.vb_labor_sga_subtotal = safe_num(value_benefit[f"T{m.value_row}"].value)
                m.vb_logistics = safe_num(value_benefit[f"U{m.value_row}"].value)

            if m.adoption_base_row is not None:
                m.adoption_sequence_axis = [int(safe_num(adoption.cell(8, c).value)) for c in range(3, 20)]
                m.adoption_recurrent_curve = [safe_num(adoption.cell(m.adoption_base_row, c).value) for c in range(3, 20)]
                m.adoption_one_time_curve = [safe_num(adoption.cell(m.adoption_base_row + 1, c).value) for c in range(3, 20)]

            if m.timeline_status_row is not None:
                m.timeline_statuses = [int(safe_num(timeline.cell(m.timeline_status_row, c).value)) for c in range(4, 24)]
                m.timeline_live_sequence = [int(safe_num(timeline.cell(m.timeline_sequence_row, c).value)) for c in range(4, 24)]
                m.financing_quarters = int(safe_num(timeline["Z10"].value))
                m.implementation_quarters = sum(1 for x in m.timeline_statuses if x == 1)

            if m.investment_row is not None and m.investment_row <= investment.max_row:
                r = m.investment_row
                m.oracle_services = safe_num(investment[f"J{r}"].value)
                m.third_party_services = safe_num(investment[f"R{r}"].value)
                m.client_services = safe_num(investment[f"Z{r}"].value)
                m.software_fee = safe_num(investment[f"AG{r}"].value)
                m.software_maintenance = safe_num(investment[f"AH{r}"].value)
                m.hardware_capital = safe_num(investment[f"AI{r}"].value)
                m.hardware_maintenance = safe_num(investment[f"AJ{r}"].value)
                m.hosting_fees = safe_num(investment[f"AK{r}"].value)

            if m.roi_start_row is not None:
                rs = m.roi_start_row
                m.sales_business_impacted_pct = safe_num(roi_modular[f"F{rs}"].value)
                m.impacted_margin_pct = safe_num(roi_modular[f"J{rs}"].value)
                m.inventory_business_impacted_pct = safe_num(roi_modular[f"F{rs + 1}"].value)
                m.labor_pct_of_sga = safe_num(roi_modular[f"J{rs + 1}"].value)

            modules.append(m)
        return modules

    def compute_forward_financials(self, fin: FinancialInputs) -> Dict[str, List[float]]:
        """
        Replicates the logic from Forward Financials.

        Revenue_qt = quarterly revenue x (1 + annual revenue growth)^(year index)
        COGS_qt = base COGS x quarterly_mix x (1 + annual cogs growth)^(year index)
        SGA_qt  = base SGA  x quarterly_mix x (1 + annual sga growth)^(year index)
        Inventory_qt = COGS_qt / annual inventory turns
        """
        revenue: List[float] = []
        cogs: List[float] = []
        gross_margin_dollars: List[float] = []
        gross_margin_pct: List[float] = []
        sga: List[float] = []
        taxes: List[float] = []
        operating_income_net_of_taxes: List[float] = []
        inventory: List[float] = []

        for i in range(TOTAL_QUARTERS):
            year_idx = i // 4
            q_idx = i % 4
            rev = fin.quarterly_revenue[q_idx] * ((1.0 + fin.annual_revenue_growth_rate) ** (year_idx + 1))
            cogs_q = fin.cogs * fin.quarterly_mix[q_idx] * ((1.0 + fin.annual_cogs_growth_rate) ** (year_idx + 1))
            gm = rev - cogs_q
            gm_pct = (gm / rev) if rev else 0.0
            sga_q = fin.sga * fin.quarterly_mix[q_idx] * ((1.0 + fin.annual_sga_growth_rate) ** (year_idx + 1))
            tax_q = (gm - sga_q) * fin.tax_rate
            op_inc_net = gm - sga_q - tax_q
            inv_q = (cogs_q / fin.annual_inventory_turns) if fin.annual_inventory_turns else 0.0

            revenue.append(rev)
            cogs.append(cogs_q)
            gross_margin_dollars.append(gm)
            gross_margin_pct.append(gm_pct)
            sga.append(sga_q)
            taxes.append(tax_q)
            operating_income_net_of_taxes.append(op_inc_net)
            inventory.append(inv_q)

        return {
            "revenue": revenue,
            "cogs": cogs,
            "gross_margin_dollars": gross_margin_dollars,
            "gross_margin_pct": gross_margin_pct,
            "sga": sga,
            "taxes": taxes,
            "operating_income_net_of_taxes": operating_income_net_of_taxes,
            "inventory": inventory,
        }

    def compute_module_quarterly(self, fin: FinancialInputs, forward: Dict[str, List[float]], module: ModuleAssumptions) -> ModuleQuarterlyResult:
        revenue = forward["revenue"]
        sga_quarterly = forward["sga"]

        profit_sales: List[float] = []
        inventory_reduction: List[float] = []
        one_time_inventory_reduction: List[float] = []
        carrying_cost_reduction: List[float] = []
        margin_improvement: List[float] = []
        labor_reduction: List[float] = []
        logistics_reduction: List[float] = []
        total_benefit: List[float] = []

        software_fees: List[float] = []
        software_maintenance: List[float] = []
        oracle_services: List[float] = []
        client_services: List[float] = []
        third_party_services: List[float] = []
        hardware_capital: List[float] = []
        hardware_maintenance: List[float] = []
        hosting_fees: List[float] = []
        total_investment: List[float] = []

        if not module.selected or module.benefit_row is None or module.roi_start_row is None:
            zeros = [0.0] * TOTAL_QUARTERS
            return ModuleQuarterlyResult(
                module_name=module.module_name,
                selected=module.selected,
                profit_from_increased_sales=zeros.copy(),
                inventory_reductions=zeros.copy(),
                one_time_inventory_reduction=zeros.copy(),
                carrying_cost_reduction=zeros.copy(),
                margin_improvement=zeros.copy(),
                labor_sga_reduction=zeros.copy(),
                logistics_reduction=zeros.copy(),
                total_estimated_benefit=zeros.copy(),
                software_fees=zeros.copy(),
                software_maintenance=zeros.copy(),
                oracle_services=zeros.copy(),
                client_services=zeros.copy(),
                third_party_integrator_services=zeros.copy(),
                hardware_capital_costs=zeros.copy(),
                hardware_maintenance=zeros.copy(),
                hosting_fees=zeros.copy(),
                total_investment=zeros.copy(),
                cumulative_estimated_benefit=zeros.copy(),
                cumulative_investment=zeros.copy(),
                cumulative_net_benefit=zeros.copy(),
                cumulative_roi=zeros.copy(),
            )

        impl_quarters = module.implementation_quarters if module.implementation_quarters > 0 else 1
        fin_quarters = module.financing_quarters if module.financing_quarters > 0 else 1

        prev_inventory_reduction = 0.0

        for i in range(TOTAL_QUARTERS):
            q_label = QUARTER_LABELS[i]
            q_factor = quarter_factor(i, fin.quarterly_mix)
            status = module.timeline_statuses[i] if i < len(module.timeline_statuses) else 0
            seq = module.timeline_live_sequence[i] if i < len(module.timeline_live_sequence) else 0

            adoption_recurrent = lookup_adoption_value(seq, module.adoption_sequence_axis, module.adoption_recurrent_curve) if seq is not None else 0.0
            adoption_one_time = lookup_adoption_value(seq, module.adoption_sequence_axis, module.adoption_one_time_curve) if seq is not None else 0.0

            # ROI Output - Modular row 11
            sales_base = (
                module.sales_business_impacted_pct
                * (
                    (revenue[i] * module.benefit_increased_sales_pct * module.impacted_margin_pct)
                    - (revenue[i] * module.benefit_increased_sales_pct * (fin.sga / fin.net_revenue if fin.net_revenue else 0.0) * module.labor_pct_of_sga)
                )
            )
            sales_benchmark = module.vb_profit_sales_subtotal * q_factor * adoption_recurrent
            profit_sales_q = adoption_recurrent * sales_base + sales_benchmark

            # ROI Output - Modular row 13
            inventory_reduction_q = (
                module.inventory_business_impacted_pct
                * fin.total_inventory
                * module.benefit_inventory_reduction_pct
                + (module.vb_inventory_reduction * q_factor)
            )

            # ROI Output - Modular row 14
            one_time_inventory_q = 0.0
            if seq > 0:
                one_time_inventory_q = (
                    module.inventory_business_impacted_pct
                    * (fin.total_inventory * module.benefit_inventory_reduction_pct + module.vb_one_time_inventory)
                    * adoption_one_time
                )

            # ROI Output - Modular row 15
            carrying_cost_q = 0.0
            if status == 2:
                inventory_base = inventory_reduction_q
                carrying_cost_q = adoption_recurrent * (
                    inventory_base * fin.inventory_carrying_cost_of_capital / 4.0
                    + module.vb_carrying_cost_benchmark * fin.inventory_carrying_cost_of_capital / 4.0
                )

            # ROI Output - Modular row 16
            margin_improvement_q = (
                adoption_recurrent
                * (
                    module.sales_business_impacted_pct
                    * revenue[i]
                    * module.benefit_margin_improvement_pct
                    * module.impacted_margin_pct
                )
                + (
                    module.vb_margin_improvement
                    * q_factor
                    * adoption_recurrent
                )
            )

            # ROI Output - Modular row 17
            labor_reduction_q = (
                adoption_recurrent
                * (module.labor_pct_of_sga * sga_quarterly[i] * module.benefit_labor_reduction_pct)
                + (module.vb_labor_sga_subtotal * q_factor * adoption_recurrent)
            )

            # ROI Output - Modular row 18
            logistics_reduction_q = (
                adoption_recurrent
                * (module.sales_business_impacted_pct * sga_quarterly[i] * module.benefit_logistics_reduction_pct)
                + (module.vb_logistics * q_factor * adoption_recurrent)
            )

            total_benefit_q = (
                profit_sales_q
                + one_time_inventory_q
                + carrying_cost_q
                + margin_improvement_q
                + labor_reduction_q
                + logistics_reduction_q
            )

            # Investment logic mirrors ROI Output - Modular rows 24:41
            financing_live = status == 2
            first_impl_quarter = status == 1 and (i == 0 or (module.timeline_statuses[i - 1] == 0 if i > 0 else True))
            impl_active = status == 1
            live_or_post = status > 0

            software_fee_q = (module.software_fee / fin_quarters) if financing_live else 0.0
            software_maint_q = (module.software_maintenance / 4.0) if module.software_fee >= 0 else 0.0
            oracle_services_q = (module.oracle_services / impl_quarters) if impl_active and module.oracle_services > 0 else 0.0
            client_services_q = (module.client_services / impl_quarters) if impl_active and module.client_services > 0 else 0.0
            third_party_services_q = (module.third_party_services / impl_quarters) if impl_active and module.third_party_services > 0 else 0.0
            hardware_capital_q = module.hardware_capital if first_impl_quarter and module.hardware_capital > 0 else 0.0
            hardware_maint_q = (module.hardware_maintenance / 4.0) if live_or_post and module.hardware_capital > 0 else 0.0
            hosting_fee_q = (module.hosting_fees / 4.0) if live_or_post and module.hosting_fees > 0 else 0.0

            total_inv_q = (
                software_fee_q
                + software_maint_q
                + oracle_services_q
                + client_services_q
                + third_party_services_q
                + hardware_capital_q
                + hardware_maint_q
                + hosting_fee_q
            )

            profit_sales.append(profit_sales_q)
            inventory_reduction.append(inventory_reduction_q)
            one_time_inventory_reduction.append(one_time_inventory_q)
            carrying_cost_reduction.append(carrying_cost_q)
            margin_improvement.append(margin_improvement_q)
            labor_reduction.append(labor_reduction_q)
            logistics_reduction.append(logistics_reduction_q)
            total_benefit.append(total_benefit_q)

            software_fees.append(software_fee_q)
            software_maintenance.append(software_maint_q)
            oracle_services.append(oracle_services_q)
            client_services.append(client_services_q)
            third_party_services.append(third_party_services_q)
            hardware_capital.append(hardware_capital_q)
            hardware_maintenance.append(hardware_maint_q)
            hosting_fees.append(hosting_fee_q)
            total_investment.append(total_inv_q)

            prev_inventory_reduction = inventory_reduction_q

        cum_benefit = cumulative(total_benefit)
        cum_investment = cumulative(total_investment)
        cum_net = [b - inv for b, inv in zip(cum_benefit, cum_investment)]
        cum_roi = [(b / inv) if inv else 0.0 for b, inv in zip(cum_benefit, cum_investment)]

        return ModuleQuarterlyResult(
            module_name=module.module_name,
            selected=module.selected,
            profit_from_increased_sales=profit_sales,
            inventory_reductions=inventory_reduction,
            one_time_inventory_reduction=one_time_inventory_reduction,
            carrying_cost_reduction=carrying_cost_reduction,
            margin_improvement=margin_improvement,
            labor_sga_reduction=labor_reduction,
            logistics_reduction=logistics_reduction,
            total_estimated_benefit=total_benefit,
            software_fees=software_fees,
            software_maintenance=software_maintenance,
            oracle_services=oracle_services,
            client_services=client_services,
            third_party_integrator_services=third_party_services,
            hardware_capital_costs=hardware_capital,
            hardware_maintenance=hardware_maintenance,
            hosting_fees=hosting_fees,
            total_investment=total_investment,
            cumulative_estimated_benefit=cum_benefit,
            cumulative_investment=cum_investment,
            cumulative_net_benefit=cum_net,
            cumulative_roi=cum_roi,
        )

    def aggregate_quarterly_total(self, module_results: List[ModuleQuarterlyResult]) -> Dict[str, List[float]]:
        def total(attr: str) -> List[float]:
            return [sum(getattr(m, attr)[i] for m in module_results if m.selected) for i in range(TOTAL_QUARTERS)]

        out = {
            "profit_from_increased_sales": total("profit_from_increased_sales"),
            "one_time_inventory_reduction": total("one_time_inventory_reduction"),
            "carrying_cost_reduction": total("carrying_cost_reduction"),
            "margin_improvement": total("margin_improvement"),
            "labor_sga_reduction": total("labor_sga_reduction"),
            "logistics_reduction": total("logistics_reduction"),
            "total_estimated_benefit": total("total_estimated_benefit"),

            "software_fees": total("software_fees"),
            "software_maintenance": total("software_maintenance"),
            "oracle_services": total("oracle_services"),
            "client_services": total("client_services"),
            "third_party_integrator_services": total("third_party_integrator_services"),
            "hardware_capital_costs": total("hardware_capital_costs"),
            "hardware_maintenance": total("hardware_maintenance"),
            "hosting_fees": total("hosting_fees"),
            "total_investment": total("total_investment"),
        }

        out["discounted_benefit"] = out["total_estimated_benefit"].copy()
        out["cumulative_estimated_benefit"] = cumulative(out["discounted_benefit"])
        out["cumulative_investment"] = cumulative(out["total_investment"])
        out["cumulative_net_benefit"] = [
            b - inv for b, inv in zip(out["cumulative_estimated_benefit"], out["cumulative_investment"])
        ]
        out["cumulative_roi"] = [
            (b / inv) if inv else 0.0
            for b, inv in zip(out["cumulative_estimated_benefit"], out["cumulative_investment"])
        ]
        return out

    def aggregate_annual_total(self, quarterly_total: Dict[str, List[float]]) -> Dict[str, List[float]]:
        annual = {k: annualize(v) for k, v in quarterly_total.items() if isinstance(v, list) and len(v) == TOTAL_QUARTERS}
        annual["cumulative_estimated_benefit"] = cumulative(annual["total_estimated_benefit"])
        annual["cumulative_investment"] = cumulative(annual["total_investment"])
        annual["cumulative_net_benefit"] = [
            b - inv for b, inv in zip(annual["cumulative_estimated_benefit"], annual["cumulative_investment"])
        ]
        annual["cumulative_roi_pct"] = [
            ((b / inv) * 100.0) if inv else 0.0
            for b, inv in zip(annual["cumulative_estimated_benefit"], annual["cumulative_investment"])
        ]
        return annual

    def compute_pnl_cashflow(self, fin: FinancialInputs, annual_total: Dict[str, List[float]]) -> Dict[str, List[float]]:
        investment_expense_non_capex = [
            annual_total["hardware_maintenance"][i]
            + annual_total["hosting_fees"][i]
            for i in range(TOTAL_YEARS)
        ]
        # Workbook keeps depreciation / amortization at 0 in the provided file.
        book_dep_hardware = [0.0] * TOTAL_YEARS
        book_amort_software_services = [0.0] * TOTAL_YEARS

        total_costs = [
            investment_expense_non_capex[i]
            + annual_total["hardware_capital_costs"][i]
            + annual_total["software_fees"][i]
            + annual_total["oracle_services"][i]
            + annual_total["client_services"][i]
            + annual_total["third_party_integrator_services"][i]
            for i in range(TOTAL_YEARS)
        ]

        incremental_ebit = [annual_total["total_estimated_benefit"][i] - total_costs[i] for i in range(TOTAL_YEARS)]
        incremental_ebitda = [
            incremental_ebit[i] + book_dep_hardware[i] + book_amort_software_services[i]
            for i in range(TOTAL_YEARS)
        ]
        one_time_inventory_savings = annual_total["carrying_cost_reduction"][:]  # mirrors workbook memo row behavior
        ebit = incremental_ebit[:]
        tax_cash = [
            (ebit[i] - book_dep_hardware[i] - book_amort_software_services[i]) * fin.tax_rate
            for i in range(TOTAL_YEARS)
        ]
        capital = [0.0] * TOTAL_YEARS
        cash_flow = [ebit[i] - tax_cash[i] - capital[i] for i in range(TOTAL_YEARS)]

        return {
            "profit_from_increased_sales": annual_total["profit_from_increased_sales"],
            "one_time_inventory_reduction": annual_total["one_time_inventory_reduction"],
            "carrying_cost_reduction": annual_total["carrying_cost_reduction"],
            "margin_improvement": annual_total["margin_improvement"],
            "labor_sga_reduction": annual_total["labor_sga_reduction"],
            "logistics_reduction": annual_total["logistics_reduction"],
            "total_estimated_benefit": annual_total["total_estimated_benefit"],
            "investment_expense_non_capex": investment_expense_non_capex,
            "book_dep_hardware": book_dep_hardware,
            "book_amort_software_services": book_amort_software_services,
            "total_costs": total_costs,
            "incremental_ebit": incremental_ebit,
            "incremental_ebitda": incremental_ebitda,
            "one_time_inventory_savings_memo": one_time_inventory_savings,
            "cash_flow": cash_flow,
        }

    def compute_proforma(self, fin: FinancialInputs, forward: Dict[str, List[float]], annual_total: Dict[str, List[float]]) -> Dict[str, Dict[str, List[float] | float]]:
        annual_forward_revenue = annualize(forward["revenue"])
        annual_forward_cogs = annualize(forward["cogs"])
        annual_forward_sga = annualize(forward["sga"])
        annual_forward_taxes = annualize(forward["taxes"])
        annual_forward_inventory = annualize(forward["inventory"])

        net_revenue = annual_forward_revenue
        cogs = annual_forward_cogs
        gross_profit = [net_revenue[i] - cogs[i] + annual_total["profit_from_increased_sales"][i] for i in range(TOTAL_YEARS)]
        operating_expenses = [
            annual_forward_sga[i]
            - annual_total["labor_sga_reduction"][i]
            - annual_total["logistics_reduction"][i]
            for i in range(TOTAL_YEARS)
        ]
        operating_income = [gross_profit[i] - operating_expenses[i] for i in range(TOTAL_YEARS)]
        taxes = annual_forward_taxes
        net_income = [operating_income[i] - taxes[i] for i in range(TOTAL_YEARS)]

        total_inventory = []
        cumulative_one_time_inventory = cumulative(annual_total["one_time_inventory_reduction"])
        for i in range(TOTAL_YEARS):
            total_inventory.append(annual_forward_inventory[i] - cumulative_one_time_inventory[i])

        balance_sheet = {
            "cash": [fin.cash] + [fin.cash] * (TOTAL_YEARS - 1),
            "accounts_receivable": [fin.accounts_receivable] + [fin.accounts_receivable] * (TOTAL_YEARS - 1),
            "total_inventory": total_inventory,
            "other_current_assets": [fin.other_current_assets] * TOTAL_YEARS,
            "property_equipment": [fin.property_equipment] * TOTAL_YEARS,
            "goodwill": [fin.goodwill] * TOTAL_YEARS,
            "other_assets": [fin.other_assets] * TOTAL_YEARS,
            "accounts_payable": [fin.accounts_payable] * TOTAL_YEARS,
            "other_current_liabilities": [fin.other_current_liabilities] * TOTAL_YEARS,
            "long_term_debt": [fin.long_term_debt] * TOTAL_YEARS,
            "other_long_term_liabilities": [fin.other_long_term_liabilities] * TOTAL_YEARS,
        }

        return {
            "income_statement": {
                "net_revenue": net_revenue,
                "cogs": cogs,
                "gross_profit": gross_profit,
                "operating_expenses": operating_expenses,
                "operating_income": operating_income,
                "taxes": taxes,
                "net_income": net_income,
            },
            "balance_sheet": balance_sheet,
        }

    def run(self) -> ModelOutputs:
        fin = self.load_financial_inputs()
        modules = self.load_modules()
        forward = self.compute_forward_financials(fin)
        module_results = [self.compute_module_quarterly(fin, forward, m) for m in modules]
        quarterly_total = self.aggregate_quarterly_total(module_results)
        annual_total = self.aggregate_annual_total(quarterly_total)
        pnl_cashflow = self.compute_pnl_cashflow(fin, annual_total)
        proforma = self.compute_proforma(fin, forward, annual_total)

        annual_cash_flow = pnl_cashflow["cash_flow"]
        summary = {
            "client_name": fin.client_name,
            "selected_modules": [m.module_name for m in modules if m.selected],
            "irr_of_cash_flows": irr(annual_cash_flow),
            "npv_of_cash_flows": npv(fin.npv_discount_rate, annual_cash_flow),
            "payback_period": payback_period(annual_total["cumulative_net_benefit"]),
        }

        return ModelOutputs(
            forward_financials=forward,
            module_results=module_results,
            quarterly_total=quarterly_total,
            annual_total=annual_total,
            pnl_cashflow=pnl_cashflow,
            proforma=proforma,
            summary_metrics=summary,
        )


def outputs_to_jsonable(outputs: ModelOutputs) -> Dict[str, Any]:
    return {
        "forward_financials": outputs.forward_financials,
        "module_results": [asdict(x) for x in outputs.module_results],
        "quarterly_total": outputs.quarterly_total,
        "annual_total": outputs.annual_total,
        "pnl_cashflow": outputs.pnl_cashflow,
        "proforma": outputs.proforma,
        "summary_metrics": outputs.summary_metrics,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Run the retail ROI model from the Excel workbook.")
    parser.add_argument("workbook", help="Path to the .xlsm workbook")
    parser.add_argument("--out", help="Path to save JSON output", default=None)
    args = parser.parse_args()

    model = WorkbookDrivenRetailROIModel(args.workbook)
    outputs = model.run()
    payload = outputs_to_jsonable(outputs)

    if args.out:
        Path(args.out).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Saved output to: {args.out}")
    else:
        print(json.dumps(payload["summary_metrics"], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
